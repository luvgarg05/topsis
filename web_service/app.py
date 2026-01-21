#!/usr/bin/env python3
"""
Flask Web Service for TOPSIS Analysis
Features: File upload, multi-format support, email delivery, real-time validation
"""

import os
import sys
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
import tempfile
import json
from datetime import datetime
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.encoders import encode_base64
import re
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)

# Configuration
if os.environ.get('VERCEL'):
    UPLOAD_FOLDER = '/tmp/uploads'
else:
    UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'json'}
MAX_FILE_SIZE = int(os.getenv('MAX_FILE_SIZE', 10 * 1024 * 1024))  # 10MB default

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# SMTP Configuration
SENDER_EMAIL = os.getenv('SENDER_EMAIL')
SENDER_PASSWORD = os.getenv('SENDER_PASSWORD')
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def convert_to_csv(filepath):
    """
    Convert Excel or JSON files to CSV format.
    
    Args:
        filepath (str): Path to the input file
    
    Returns:
        pd.DataFrame: Converted data as DataFrame
    """
    ext = filepath.rsplit('.', 1)[1].lower()
    
    try:
        if ext == 'csv':
            return pd.read_csv(filepath)
        
        elif ext in ['xlsx', 'xls']:
            # Use data_only=True to read calculated values instead of formulas
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            # Create DataFrame from data
            if data:
                df = pd.DataFrame(data[1:], columns=data[0])
                return df
            else:
                raise ValueError("Empty Excel file")
        
        elif ext == 'json':
            # Handle both list of objects and direct data
            with open(filepath, 'r') as f:
                data = json.load(f)
            
            if isinstance(data, list):
                return pd.DataFrame(data)
            else:
                return pd.DataFrame([data])
        
        else:
            raise ValueError(f"Unsupported file format: {ext}")
    
    except Exception as e:
        raise ValueError(f"Error converting file: {str(e)}")


def validate_email(email):
    """Validate email format."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def parse_weights(weights_str):
    """Parse and validate weights."""
    try:
        weights = [float(w.strip()) for w in weights_str.split(',') if w.strip()]
        if not weights:
            return None, "Weights cannot be empty"
        if any(w <= 0 for w in weights):
            return None, "All weights must be positive numbers (> 0)"
        return weights, None
    except ValueError:
        return None, "Weights must be valid numbers. Use: 1,2,3 (comma-separated)"


def parse_impacts(impacts_str):
    """Parse and validate impacts."""
    try:
        # Check if user forgot commas (e.g., "+-" instead of "+,-")
        if ',' not in impacts_str and any(c in impacts_str for c in ['+', '-']):
            # Try to auto-fix by adding commas between symbols
            fixed = ','.join(impacts_str.replace(' ', ''))
            impacts_str = fixed
        
        impacts = [i.strip() for i in impacts_str.split(',')]
        # Filter out empty strings
        impacts = [i for i in impacts if i]
        if not impacts:
            return None, "Impacts cannot be empty"
        if not all(i in ['+', '-'] for i in impacts):
            invalid = [i for i in impacts if i not in ['+', '-']]
            return None, f"Invalid impacts: {invalid}. Use only '+' (benefit) or '-' (cost). Example: +,+,-,+"
        return impacts, None
    except Exception as e:
        return None, "Invalid impacts format. Use: +,+,- (comma-separated, like: +,-,+,+)"


def validate_inputs(df, weights, impacts):
    """Validate TOPSIS inputs."""
    # Requirement: Input file must contain three or more columns
    if len(df.columns) < 3:
        return False, "Input data must contain three or more columns"
    
    try:
        pd.to_numeric(df.iloc[:, 0])
        return False, "First column must contain non-numeric identifiers"
    except (ValueError, TypeError):
        pass
    
    num_criteria = len(df.columns) - 1
    
    if len(weights) != num_criteria:
        return False, f"Number of weights ({len(weights)}) must match criteria ({num_criteria})"
    
    if len(impacts) != num_criteria:
        return False, f"Number of impacts ({len(impacts)}) must match criteria ({num_criteria})"
    
    criteria_data = df.iloc[:, 1:]
    try:
        numeric_df = criteria_data.apply(pd.to_numeric, errors='raise')
    except (ValueError, TypeError):
        return False, "All criteria must be numeric"
    
    if (numeric_df <= 0).any().any():
        return False, "All criteria values must be positive"
    
    return True, "Valid"


def normalize_matrix(df):
    """Normalize the decision matrix."""
    matrix = df.values.astype(float)
    norms = np.sqrt((matrix ** 2).sum(axis=0))
    normalized = matrix / norms
    return normalized


def apply_weights(normalized_matrix, weights):
    """Apply weights to normalized matrix."""
    return normalized_matrix * np.array(weights)


def find_ideal_solutions(weighted_matrix, impacts):
    """Find ideal and anti-ideal solutions."""
    ideal = np.zeros(weighted_matrix.shape[1])
    anti_ideal = np.zeros(weighted_matrix.shape[1])
    
    for j in range(weighted_matrix.shape[1]):
        if impacts[j] == '+':
            ideal[j] = weighted_matrix[:, j].max()
            anti_ideal[j] = weighted_matrix[:, j].min()
        else:
            ideal[j] = weighted_matrix[:, j].min()
            anti_ideal[j] = weighted_matrix[:, j].max()
    
    return ideal, anti_ideal


def calculate_separations(weighted_matrix, ideal, anti_ideal):
    """Calculate separation measures."""
    s_plus = np.sqrt(((weighted_matrix - ideal) ** 2).sum(axis=1))
    s_minus = np.sqrt(((weighted_matrix - anti_ideal) ** 2).sum(axis=1))
    return s_plus, s_minus


def calculate_scores(s_plus, s_minus):
    """Calculate TOPSIS scores and ranks."""
    with np.errstate(divide='ignore', invalid='ignore'):
        scores = np.where(s_plus + s_minus != 0, s_minus / (s_plus + s_minus), 0)
    
    ranks = pd.Series(scores).rank(method='min', ascending=False).astype(int).values
    return scores, ranks


def perform_topsis(df, weights, impacts):
    """Execute TOPSIS analysis."""
    criteria = df.iloc[:, 1:]
    
    normalized = normalize_matrix(criteria)
    weighted = apply_weights(normalized, weights)
    ideal, anti_ideal = find_ideal_solutions(weighted, impacts)
    s_plus, s_minus = calculate_separations(weighted, ideal, anti_ideal)
    scores, ranks = calculate_scores(s_plus, s_minus)
    
    result_df = df.copy()
    result_df['Topsis Score'] = scores
    result_df['Rank'] = ranks
    result_df = result_df.sort_values('Rank')
    
    return result_df


def send_email(recipient_email, subject, body, attachment_path=None):
    """Send email with optional attachment."""
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        print(f"[EMAIL] Not configured: SENDER_EMAIL={SENDER_EMAIL}, PASSWORD={'***' if SENDER_PASSWORD else 'NOT SET'}")
        return False, "Email not configured. Check .env file."
    
    if "your-app-specific-password" in SENDER_PASSWORD:
        print(f"[EMAIL] Password is placeholder")
        return False, "Email password not set. Use actual app-specific password in .env"
    
    try:
        print(f"[EMAIL] Preparing email to {recipient_email}")
        message = MIMEMultipart()
        message['From'] = SENDER_EMAIL
        message['To'] = recipient_email
        message['Subject'] = subject
        
        message.attach(MIMEText(body, 'html'))
        
        if attachment_path and os.path.exists(attachment_path):
            print(f"[EMAIL] Attaching file: {attachment_path}")
            with open(attachment_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment_path)}')
            message.attach(part)
        
        print(f"[EMAIL] Connecting to {SMTP_SERVER}:{SMTP_PORT}")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            print(f"[EMAIL] Logging in as {SENDER_EMAIL}")
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            print(f"[EMAIL] Sending message...")
            server.send_message(message)
        
        print(f"[EMAIL] Email sent successfully to {recipient_email}")
        return True, "Email sent successfully"
    
    except smtplib.SMTPAuthenticationError as e:
        msg = f"Email authentication failed. Check credentials in .env file."
        print(f"[EMAIL] Error: {msg}")
        return False, msg
    except smtplib.SMTPException as e:
        msg = f"Email SMTP error: {str(e)}"
        print(f"[EMAIL] Error: {msg}")
        return False, msg
    except Exception as e:
        msg = f"Email error: {str(e)}"
        print(f"[EMAIL] Error: {msg}")
        return False, msg


@app.route('/')
def index():
    """Serve the main page."""
    return render_template('index.html')


@app.route('/api/validate', methods=['POST'])
def validate():
    """Validate user inputs."""
    try:
        data = request.json
        email = data.get('email', '')
        weights_str = data.get('weights', '')
        impacts_str = data.get('impacts', '')
        
        errors = {}
        
        # Validate email
        if not email:
            errors['email'] = 'Email is required'
        elif not validate_email(email):
            errors['email'] = 'Invalid email format'
        
        # Validate weights
        if not weights_str:
            errors['weights'] = 'Weights are required'
        else:
            weights, err = parse_weights(weights_str)
            if err:
                errors['weights'] = err
        
        # Validate impacts
        if not impacts_str:
            errors['impacts'] = 'Impacts are required'
        else:
            impacts, err = parse_impacts(impacts_str)
            if err:
                errors['impacts'] = err
        
        return jsonify({'valid': len(errors) == 0, 'errors': errors})
    
    except Exception as e:
        return jsonify({'valid': False, 'errors': {'general': str(e)}}), 400


@app.route('/api/analyze', methods=['POST'])
def analyze():
    """Analyze uploaded file using TOPSIS."""
    try:
        # Check required fields
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        email = request.form.get('email')
        weights_str = request.form.get('weights')
        impacts_str = request.form.get('impacts')
        
        if not all([email, weights_str, impacts_str]):
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        # Validate email
        if not validate_email(email):
            return jsonify({'success': False, 'error': 'Invalid email format'}), 400
        
        # Parse and validate weights
        weights, err = parse_weights(weights_str)
        if err:
            return jsonify({'success': False, 'error': f'Weights error: {err}'}), 400
        
        # Parse and validate impacts
        impacts, err = parse_impacts(impacts_str)
        if err:
            return jsonify({'success': False, 'error': f'Impacts error: {err}'}), 400
        
        # Handle file upload
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'File type not supported'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Convert file to DataFrame
            df = convert_to_csv(filepath)
            
            # Validate inputs
            is_valid, message = validate_inputs(df, weights, impacts)
            if not is_valid:
                return jsonify({'success': False, 'error': message}), 400
            
            # Perform TOPSIS analysis
            result_df = perform_topsis(df, weights, impacts)
            
            # Save results
            result_filename = f"topsis_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
            result_df.to_csv(result_path, index=False)
            
            # Prepare response with full data
            results = []
            full_data = []
            for idx, row in result_df.iterrows():
                result_entry = {}
                # Add all columns from the original data
                for col_idx, col in enumerate(result_df.columns):
                    if col == 'Topsis Score':
                        result_entry['Topsis Score'] = round(float(row[col]), 6)
                    elif col == 'Rank':
                        result_entry['Rank'] = int(row[col])
                    else:
                        # Try to convert to number, keep as string if fails
                        try:
                            result_entry[col] = float(row[col])
                        except:
                            result_entry[col] = str(row[col])
                
                full_data.append(result_entry)
                
                # For simple results list (used by ranking display)
                results.append({
                    'name': row.iloc[0],
                    'score': round(float(row['Topsis Score']), 4),
                    'rank': int(row['Rank'])
                })
            
            # Send email
            email_body = f"""
            <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2>TOPSIS Analysis Results</h2>
                    <p>Hi,</p>
                    <p>Your TOPSIS analysis has been completed successfully!</p>
                    <h3>Analysis Summary</h3>
                    <ul>
                        <li><strong>Criteria:</strong> {len(df.columns) - 1}</li>
                        <li><strong>Alternatives:</strong> {len(df)}</li>
                        <li><strong>Weights:</strong> {', '.join(map(str, weights))}</li>
                        <li><strong>Impacts:</strong> {', '.join(impacts)}</li>
                    </ul>
                    <h3>Top Results</h3>
                    <table border="1" cellpadding="10">
                        <tr>
                            <th>Rank</th>
                            <th>Alternative</th>
                            <th>Score</th>
                        </tr>
            """
            
            for result in results[:5]:
                email_body += f"""
                        <tr>
                            <td>{result['rank']}</td>
                            <td>{result['name']}</td>
                            <td>{result['score']}</td>
                        </tr>
                """
            
            email_body += """
                    </table>
                    <p style="margin-top: 20px; color: #666;">
                        Detailed results are attached to this email.
                    </p>
                </body>
            </html>
            """
            
            success, msg = send_email(email, "TOPSIS Analysis Results", email_body, result_path)
            
            return jsonify({
                'success': True,
                'message': 'Analysis completed successfully!',
                'results': results,
                'full_data': full_data,
                'columns': list(result_df.columns),
                'download_filename': result_filename,
                'email_sent': success
            })
        
        finally:
            # Clean up uploaded file
            if os.path.exists(filepath):
                os.remove(filepath)
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download/<filename>')
def download(filename):
    """Download result file."""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.errorhandler(413)
def file_too_large(e):
    """Handle file too large error."""
    return jsonify({'error': 'File too large. Maximum size: 10MB'}), 413


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
